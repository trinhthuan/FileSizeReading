import sys
import os

from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QFileDialog, QTableWidget, \
    QTableWidgetItem, QLabel, QCheckBox, QProgressBar, QMessageBox, QHeaderView

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment



class FileSizeChecker(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("FolderSizeMeasure - TrinhThuan - E20250308")
        self.setGeometry(100, 100, 600, 500)
        self.setMinimumSize(200, 150)  # Cho phép kéo nhỏ hơn kích thước ban đầu

        self.setStyleSheet("background-color: #f0f0f0;")
        # Kích hoạt kéo thả
        self.setAcceptDrops(True)

        layout = QVBoxLayout()

        self.label = QLabel("Chọn thư mục để kiểm tra dung lượng các file")
        self.label.setStyleSheet("font-size: 14px; font-weight: bold;")
        layout.addWidget(self.label)

        button_layout = QHBoxLayout()

        self.button = QPushButton("Chọn thư mục")
        self.button.setStyleSheet(
            "background-color: #4CAF50; color: white; font-size: 14px; padding: 8px; border-radius: 5px;")
        self.button.clicked.connect(self.select_folder)
        button_layout.addWidget(self.button)

        self.export_button = QPushButton("Xuất Excel")
        self.export_button.setStyleSheet(
            "background-color: #2196F3; color: white; font-size: 14px; padding: 8px; border-radius: 5px;")
        self.export_button.clicked.connect(self.export_to_excel)
        button_layout.addWidget(self.export_button)

        layout.addLayout(button_layout)

        self.detail_checkbox = QCheckBox("Xem chi tiết tất cả file và thư mục con")
        layout.addWidget(self.detail_checkbox)



        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Tên file", "Dung lượng", "Đơn vị"])
        self.table.setStyleSheet("background-color: white; border: 1px solid #ccc;")

        # Cột đầu tiên mở rộng theo bảng
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)

        # Các cột khác chỉ điều chỉnh theo nội dung
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

        # self.table.horizontalHeader().setStretchLastSection(True)  # Giãn cột cuối cùng để không bị thừa khoảng trống
        # self.table.setColumnWidth(0, 400)  # Tăng chiều rộng cột tên file
        # self.table.setColumnWidth(1, 100)
        # self.table.setColumnWidth(2, 100)
        layout.addWidget(self.table)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(10)  # Đặt chiều cao thanh tiến trình
        # self.progress_bar.setFixedWidth(600)  # Đặt chiều rộng thanh tiến trình
        self.progress_bar.setStyleSheet(
            "QProgressBar { border: 0px solid grey; border-radius: 5px; text-align: center; height: 10px; width: 500px; } "
            "QProgressBar::chunk { background-color: #00FFCC; width: 10px; }")
        layout.addWidget(self.progress_bar)

        self.total_label = QLabel("Tổng dung lượng: 0 KB")
        self.total_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #d32f2f;")
        layout.addWidget(self.total_label)

        self.setLayout(layout)

    def dragEnterEvent(self, event):
        """ Khi người dùng kéo một thư mục vào giao diện """
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        """ Khi người dùng thả thư mục vào giao diện """
        urls = event.mimeData().urls()
        if urls:
            folder_path = urls[0].toLocalFile()
            if os.path.isdir(folder_path):  # Kiểm tra xem có phải thư mục không
                self.label.setText(f"Thư mục đã chọn: {folder_path}")
                self.list_files(folder_path)

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục")
        if folder:
            self.label.setText(f"Thư mục đã chọn: {folder}")
            self.list_files(folder)

    def get_folder_size(self, folder):
        total_size = 0
        for dirpath, dirnames, filenames in os.walk(folder):
            for f in filenames:
                fp = os.path.join(dirpath, f)
                if os.path.exists(fp):
                    total_size += os.path.getsize(fp)
        return total_size

    def add_file_data(self, file, size_bytes):
        if size_bytes >= 1024 * 1024:
            size_display = size_bytes / (1024 * 1024)
            unit = "MB"
        elif size_bytes >= 1024:
            size_display = size_bytes / 1024
            unit = "KB"
        else:
            size_display = size_bytes
            unit = "Bytes"

        self.file_data.append((file, size_display, unit, size_bytes))

    def list_files(self, folder):
        self.folder = folder
        self.file_data = []
        total_size = 0

        files_list = []
        if self.detail_checkbox.isChecked():
            for dirpath, _, filenames in os.walk(folder):
                for file in filenames:
                    files_list.append(os.path.join(dirpath, file))
        else:
            files_list = [os.path.join(folder, file) for file in os.listdir(folder)]

        self.progress_bar.setMaximum(len(files_list))

        for index, file_path in enumerate(files_list):
            if os.path.isfile(file_path):
                size_bytes = os.path.getsize(file_path)
                total_size += size_bytes
            elif os.path.isdir(file_path):
                size_bytes = self.get_folder_size(file_path)
                total_size += size_bytes
            else:
                size_bytes = 0

            self.add_file_data(file_path, size_bytes)
            self.progress_bar.setValue(index + 1)

        self.file_data.sort(key=lambda x: x[3], reverse=True)

        self.table.setRowCount(len(self.file_data))
        for i, (file, size_display, unit, _) in enumerate(self.file_data):
            self.table.setItem(i, 0, QTableWidgetItem(file))

            item_size = QTableWidgetItem(f"{size_display:.2f}" if isinstance(size_display, float) else size_display)
            item_size.setTextAlignment(0x0002 | 0x0080)
            self.table.setItem(i, 1, item_size)

            item_unit = QTableWidgetItem(unit)
            item_unit.setTextAlignment(0x0004 | 0x0080)
            self.table.setItem(i, 2, item_unit)

        total_size_display, total_unit = (total_size / (1024 * 1024), "MB") if total_size >= 1024 * 1024 else (
        total_size / 1024, "KB") if total_size >= 1024 else (total_size, "Bytes")
        self.total_label.setText(f"Tổng dung lượng: {total_size_display:.2f} {total_unit}")
        self.progress_bar.setValue(len(files_list))

    def export_to_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Lưu file", "", "Excel Files (*.xlsx)")
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                data = [["Tên file", "Dung lượng", "Đơn vị"]]

                # Lấy dữ liệu từ bảng giao diện
                for row in range(self.table.rowCount()):
                    file_name = self.table.item(row, 0).text()
                    size = self.table.item(row, 1).text()
                    unit = self.table.item(row, 2).text()
                    data.append([file_name, size, unit])

                # Lấy tổng dung lượng hiển thị trên nhãn
                total_size_text = self.total_label.text().replace("Tổng dung lượng: ", "").split(" ")
                total_size, total_unit = total_size_text[0], total_size_text[1]
                data.append(["Tổng cộng", total_size, total_unit])

                # Ghi toàn bộ dữ liệu một lần
                for row in data:
                    ws.append(row)

                # Áp dụng định dạng sau khi ghi dữ liệu
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                for col in ws.columns:
                    for cell in col:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        if cell.row == 1:
                            cell.font = Font(bold=True)
                            cell.fill = yellow_fill

                wb.save(file_path)
                QMessageBox.information(self, "Xuất Excel", f"Xuất file Excel hoàn tất: {file_path}")
            except PermissionError:
                QMessageBox.warning(self, "Lỗi", "File Excel đang mở, vui lòng đóng file trước khi xuất.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FileSizeChecker()
    window.show()
    sys.exit(app.exec())
